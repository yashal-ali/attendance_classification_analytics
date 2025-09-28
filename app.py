import streamlit as st
import pandas as pd
import io
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

# Color mapping for attendance types
COLOR_MAP = {
    'present': '90EE90',  # Light green
    'outstation_present': '87CEEB',  # Light blue
    'late': 'FFB6C1',  # Light red
    'outstation_late': 'FFFACD',  # Light yellow
    'absent': 'F5F5F5'  # Light gray
}

def classify_attendance(row):
    """
    Classify attendance based on the rules:
    - Present if: Start time ≤ 9:15 AM AND Start meter < 200
    - Outstation Present if: present AND reason = outstation
    - Outstation Late if: late AND reason = outstation
    - Late: otherwise
    """
    try:
        # Extract start time and convert to datetime
        start_time = pd.to_datetime(row['Start Day Time'])
        cutoff_time = start_time.replace(hour=9, minute=15, second=0, microsecond=0)
       
        # Check start meter - handle both numeric and "Other Location"
        start_meter = row['Start DiffIn Meters']
        if isinstance(start_meter, str) and start_meter == "Other Location":
            start_meter_value = float('inf')  # Treat "Other Location" as very large distance
        else:
            try:
                start_meter_value = float(start_meter)
            except (ValueError, TypeError):
                start_meter_value = float('inf')
       
        # Check attendance reason
        reason_start = str(row['Attendance Reason Start']).lower() if pd.notna(row['Attendance Reason Start']) else ""
        reason_end = str(row['Attendance Reason End']).lower() if pd.notna(row['Attendance Reason End']) else ""
        is_outstation = "outstation" in reason_start or "outstation" in reason_end
       
        # Apply classification rules
        if start_time <= cutoff_time and start_meter_value < 200:
            if is_outstation:
                return "outstation_present"
            else:
                return "present"
        else:
            if is_outstation:
                return "outstation_late"
            else:
                return "late"
               
    except Exception as e:
        return "late"  # Default to late if there's any error

def create_monthly_calendar_view(df):
    """
    Create a monthly calendar view with User Code, User Name, and columns for each date with day name
    """
    # Extract date from Attendance Date
    df['date'] = pd.to_datetime(df['Attendance Date']).dt.date
    df['date_str'] = pd.to_datetime(df['Attendance Date']).dt.strftime('%Y-%m-%d')
    df['day_name'] = pd.to_datetime(df['Attendance Date']).dt.strftime('%a')  # Mon, Tue, etc.
    df['day_number'] = pd.to_datetime(df['Attendance Date']).dt.day
   
    # Create a combined column for display: "DD Mon" (e.g., "01 Mon")
    df['date_display'] = pd.to_datetime(df['Attendance Date']).dt.strftime('%d %a')
   
    # Get unique users and dates
    users = df[['User Code', 'User Name']].drop_duplicates().sort_values('User Code')
    unique_dates = sorted(df['date'].unique())
   
    # Create a complete grid of all users and all dates in the month
    all_dates = sorted(df['date_str'].unique())
    date_display_map = dict(zip(df['date_str'], df['date_display']))
   
    # Create pivot table using date_str as columns
    calendar_df = df.pivot_table(
        index=['User Code', 'User Name'],
        columns='date_str',
        values='attendance',
        aggfunc='first',  # Use first occurrence if multiple entries per user per day
        fill_value='absent'  # Fill missing days with 'absent'
    ).reset_index()
   
    # Ensure all date columns are present
    for date_str in all_dates:
        if date_str not in calendar_df.columns:
            calendar_df[date_str] = 'absent'
   
    # Reorder columns to have User Code, User Name, then all dates in order
    date_columns = sorted([col for col in calendar_df.columns if col not in ['User Code', 'User Name']])
    calendar_df = calendar_df[['User Code', 'User Name'] + date_columns]
   
    # Rename columns to show date display format instead of date string
    column_mapping = {'User Code': 'User Code', 'User Name': 'User Name'}
    for col in calendar_df.columns[2:]:
        if col in date_display_map:
            column_mapping[col] = date_display_map[col]
        else:
            # Fallback: convert date string to display format
            try:
                display_date = pd.to_datetime(col).strftime('%d %a')
                column_mapping[col] = display_date
            except:
                column_mapping[col] = col
   
    calendar_df = calendar_df.rename(columns=column_mapping)
   
    return calendar_df

def create_employee_summary(df):
    """
    Create employee summary table with counts of each attendance type
    """
    # Group by user and count each attendance type
    summary = df.groupby(['User Code', 'User Name'])['attendance'].value_counts().unstack(fill_value=0)
   
    # Ensure all attendance types are present as columns
    attendance_types = ['present', 'outstation_present', 'late', 'outstation_late', 'absent']
    for att_type in attendance_types:
        if att_type not in summary.columns:
            summary[att_type] = 0
   
    # Reorder columns and reset index
    summary = summary[attendance_types].reset_index()
   
    # Calculate total days
    summary['total_days'] = summary[attendance_types].sum(axis=1)
   
    # Calculate total present (present + outstation_present)
    summary['total_present'] = summary['present'] + summary['outstation_present']
   
    # Calculate attendance percentage
    summary['attendance_percentage'] = (summary['total_present'] / summary['total_days'] * 100).round(1)
   
    # Rename columns for better display
    summary.columns = ['User Code', 'User Name', 'Present', 'Outstation Present', 'Late', 'Outstation Late', 'Absent', 'Total Days', 'Total Present', 'Attendance %']
   
    # Reorder columns to have Total Present after Outstation Present
    summary = summary[['User Code', 'User Name', 'Present', 'Outstation Present', 'Total Present', 'Late', 'Outstation Late', 'Absent', 'Total Days', 'Attendance %']]
   
    return summary

def create_total_present_summary(df):
    """
    Create a simplified table with User Code, User Name and Total Present count
    """
    # Group by user and count each attendance type
    summary = df.groupby(['User Code', 'User Name'])['attendance'].value_counts().unstack(fill_value=0)
   
    # Ensure all attendance types are present as columns
    attendance_types = ['present', 'outstation_present']
    for att_type in attendance_types:
        if att_type not in summary.columns:
            summary[att_type] = 0
   
    # Calculate total present (present + outstation_present)
    summary['total_present'] = summary['present'] + summary['outstation_present']
   
    # Reset index and select only required columns
    total_present_df = summary[['total_present']].reset_index()
   
    # Rename columns for better display
    total_present_df.columns = ['User Code', 'User Name', 'Total Present']
   
    # Sort by Total Present in descending order
    total_present_df = total_present_df.sort_values('Total Present', ascending=False)
   
    return total_present_df

def create_dashboard_metrics(df, employee_summary):
    """
    Create comprehensive dashboard metrics and visualizations
    """
    # Calculate overall metrics
    total_employees = len(employee_summary)
    total_present_days = employee_summary['Total Present'].sum()
    total_late_days = employee_summary['Late'].sum() + employee_summary['Outstation Late'].sum()
    total_absent_days = employee_summary['Absent'].sum()
    overall_attendance_rate = (total_present_days / (total_employees * employee_summary['Total Days'].max()) * 100).round(1)
   
    # Daily attendance trends
    df['date'] = pd.to_datetime(df['Attendance Date']).dt.date
    daily_attendance = df.groupby('date')['attendance'].value_counts().unstack(fill_value=0)
   
    # Employee performance categories
    employee_summary['performance_category'] = pd.cut(
        employee_summary['Attendance %'],
        bins=[0, 50, 75, 90, 100],
        labels=['Poor (<50%)', 'Average (50-75%)', 'Good (75-90%)', 'Excellent (>90%)']
    )
   
    return {
        'total_employees': total_employees,
        'total_present_days': total_present_days,
        'total_late_days': total_late_days,
        'total_absent_days': total_absent_days,
        'overall_attendance_rate': overall_attendance_rate,
        'daily_attendance': daily_attendance,
        'performance_categories': employee_summary['performance_category'].value_counts()
    }

def create_visualizations(df, employee_summary, dashboard_metrics):
    """
    Create various visualizations for the dashboard
    """
    visualizations = {}
   
    # 1. Overall Attendance Distribution Pie Chart
    attendance_counts = df['attendance'].value_counts()
    fig_pie = px.pie(
        values=attendance_counts.values,
        names=attendance_counts.index,
        title='Overall Attendance Distribution',
        color=attendance_counts.index,
        color_discrete_map={
            'present': '#90EE90',
            'outstation_present': '#87CEEB',
            'late': '#FFB6C1',
            'outstation_late': '#FFFACD',
            'absent': '#F5F5F5'
        }
    )
    fig_pie.update_traces(textposition='inside', textinfo='percent+label')
    visualizations['attendance_pie'] = fig_pie
   
    # 2. Performance Categories Bar Chart
    perf_counts = dashboard_metrics['performance_categories']
    fig_perf = px.bar(
        x=perf_counts.index,
        y=perf_counts.values,
        title='Employee Performance Categories',
        labels={'x': 'Performance Category', 'y': 'Number of Employees'},
        color=perf_counts.index,
        color_discrete_sequence=px.colors.qualitative.Set3
    )
    visualizations['performance_bar'] = fig_perf
   
    # 3. Top 10 Employees by Attendance
    top_employees = employee_summary.nlargest(10, 'Attendance %')[['User Name', 'Attendance %']]
    fig_top = px.bar(
        top_employees,
        x='User Name',
        y='Attendance %',
        title='Top 10 Employees by Attendance Rate',
        color='Attendance %',
        color_continuous_scale='Viridis'
    )
    visualizations['top_employees'] = fig_top
   
    # 4. Daily Attendance Trends
    daily_data = dashboard_metrics['daily_attendance']
    fig_daily = go.Figure()
   
    colors = {
        'present': '#90EE90',
        'outstation_present': '#87CEEB',
        'late': '#FFB6C1',
        'outstation_late': '#FFFACD',
        'absent': '#F5F5F5'
    }
   
    for status in daily_data.columns:
        fig_daily.add_trace(go.Scatter(
            x=daily_data.index,
            y=daily_data[status],
            name=status,
            stackgroup='one',
            line=dict(width=0.5, color=colors.get(status, '#000000')),
            fillcolor=colors.get(status, '#000000')
        ))
   
    fig_daily.update_layout(
        title='Daily Attendance Trends (Stacked Area)',
        xaxis_title='Date',
        yaxis_title='Number of Employees'
    )
    visualizations['daily_trends'] = fig_daily
   
    # 5. Attendance Rate by Employee (Scatter Plot)
    fig_scatter = px.scatter(
        employee_summary,
        x='Total Present',
        y='Attendance %',
        size='Total Days',
        color='Attendance %',
        hover_data=['User Name'],
        title='Employee Attendance: Present Days vs Attendance Rate',
        color_continuous_scale='RdYlGn'
    )
    visualizations['attendance_scatter'] = fig_scatter
   
    return visualizations

def create_individual_analysis(df, employee_summary, selected_employee):
    """
    Create individual employee analysis with detailed breakdown and visualizations
    """
    if selected_employee == "Select Employee":
        return None
   
    # Get employee data
    if selected_employee in employee_summary['User Code'].values:
        employee_data = employee_summary[employee_summary['User Code'] == selected_employee].iloc[0]
        user_code = selected_employee
        user_name = employee_data['User Name']
    else:
        employee_data = employee_summary[employee_summary['User Name'] == selected_employee].iloc[0]
        user_code = employee_data['User Code']
        user_name = selected_employee
   
    # Filter individual records
    individual_df = df[(df['User Code'] == user_code) | (df['User Name'] == user_name)].copy()
   
    # Create detailed breakdown
    individual_df['date'] = pd.to_datetime(individual_df['Attendance Date']).dt.date
    individual_df['day_name'] = pd.to_datetime(individual_df['Attendance Date']).dt.strftime('%A')
   
    # Sort by date
    individual_df = individual_df.sort_values('date')
   
    # Create visualizations for individual analysis
    visualizations = {}
   
    # 1. Attendance Pattern by Day of Week
    day_pattern = individual_df.groupby('day_name')['attendance'].value_counts().unstack(fill_value=0)
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_pattern = day_pattern.reindex(day_order, fill_value=0)
   
    fig_day = px.bar(
        day_pattern,
        title=f'Attendance Pattern by Day of Week - {user_name}',
        labels={'value': 'Count', 'day_name': 'Day of Week'},
        barmode='stack',
        color_discrete_map={
            'present': '#90EE90',
            'outstation_present': '#87CEEB',
            'late': '#FFB6C1',
            'outstation_late': '#FFFACD',
            'absent': '#F5F5F5'
        }
    )
    fig_day.update_layout(xaxis_title='Day of Week', yaxis_title='Number of Days')
    visualizations['day_pattern'] = fig_day
   
    # 2. Monthly Attendance Trend
    individual_df['month_week'] = individual_df['date'].apply(lambda x: f"Week {(x.day-1)//7 + 1}")
    weekly_trend = individual_df.groupby('month_week')['attendance'].value_counts().unstack(fill_value=0)
   
    fig_weekly = px.line(
        weekly_trend,
        title=f'Weekly Attendance Trend - {user_name}',
        markers=True,
        labels={'value': 'Count', 'month_week': 'Week of Month'}
    )
    fig_weekly.update_layout(xaxis_title='Week of Month', yaxis_title='Number of Days')
    visualizations['weekly_trend'] = fig_weekly
   
    # 3. Attendance Distribution Pie Chart
    individual_counts = individual_df['attendance'].value_counts()
    fig_individual_pie = px.pie(
        values=individual_counts.values,
        names=individual_counts.index,
        title=f'Attendance Distribution - {user_name}',
        color=individual_counts.index,
        color_discrete_map={
            'present': '#90EE90',
            'outstation_present': '#87CEEB',
            'late': '#FFB6C1',
            'outstation_late': '#FFFACD',
            'absent': '#F5F5F5'
        }
    )
    fig_individual_pie.update_traces(textposition='inside', textinfo='percent+label')
    visualizations['individual_pie'] = fig_individual_pie
   
    # 4. Start Time Analysis (if available)
    if 'Start Day Time' in individual_df.columns:
        individual_df['start_hour'] = pd.to_datetime(individual_df['Start Day Time']).dt.hour
        individual_df['start_minute'] = pd.to_datetime(individual_df['Start Day Time']).dt.minute
        individual_df['start_time_decimal'] = individual_df['start_hour'] + individual_df['start_minute']/60
       
        fig_time = px.box(
            individual_df,
            x='attendance',
            y='start_time_decimal',
            title=f'Start Time Distribution by Attendance Type - {user_name}',
            labels={'start_time_decimal': 'Start Time (24h)', 'attendance': 'Attendance Type'},
            color='attendance',
            color_discrete_map={
                'present': '#90EE90',
                'outstation_present': '#87CEEB',
                'late': '#FFB6C1',
                'outstation_late': '#FFFACD'
            }
        )
        fig_time.update_layout(yaxis_tickformat = '.1f')
        fig_time.add_hline(y=9.25, line_dash="dash", line_color="red", annotation_text="9:15 AM Cutoff")
        visualizations['time_analysis'] = fig_time
   
    return {
        'user_code': user_code,
        'user_name': user_name,
        'summary': employee_data,
        'detailed_records': individual_df,
        'visualizations': visualizations
    }

def style_calendar(df):
    """
    Apply styling to the calendar view based on attendance status
    """
    def color_cell(val):
        if val == 'present':
            return 'background-color: #90EE90; color: #000000; font-weight: bold'
        elif val == 'outstation_present':
            return 'background-color: #87CEEB; color: #000000; font-weight: bold'
        elif val == 'late':
            return 'background-color: #FFB6C1; color: #000000; font-weight: bold'
        elif val == 'outstation_late':
            return 'background-color: #FFFACD; color: #000000; font-weight: bold'
        elif val == 'absent':
            return 'background-color: #F5F5F5; color: #999999; font-weight: bold'
        else:
            return ''
   
    # Apply styling only to date columns (columns that look like "01 Mon", "02 Tue", etc.)
    date_columns = [col for col in df.columns if any(day in col for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'])]
    styled_df = df.style.applymap(color_cell, subset=date_columns)
   
    return styled_df

def style_summary(df):
    """
    Apply styling to the summary table with improved UI
    """
    def highlight_attendance(row):
        styles = [''] * len(row)
        
        # Highlight high attendance percentage
        if row.name == 'Attendance %' and row['Attendance %'] >= 90:
            styles[row.index.get_loc('Attendance %')] = 'background-color: #90EE90; color: #000000; font-weight: bold'
        elif row.name == 'Attendance %' and row['Attendance %'] < 70:
            styles[row.index.get_loc('Attendance %')] = 'background-color: #FFB6C1; color: #000000; font-weight: bold'
            
        # Highlight high present counts
        if row.name == 'Total Present' and row['Total Present'] == df['Total Present'].max():
            styles[row.index.get_loc('Total Present')] = 'background-color: #E8F5E8; color: #000000; font-weight: bold'
            
        # Highlight high absent counts
        if row.name == 'Absent' and row['Absent'] == df['Absent'].max():
            styles[row.index.get_loc('Absent')] = 'background-color: #FFE6E6; color: #000000; font-weight: bold'
            
        return styles
    
    styled_df = df.style.apply(highlight_attendance, axis=1)
    
    # Format percentage column
    styled_df = styled_df.format({'Attendance %': '{:.1f}%'})
    
    return styled_df

def style_total_present(df):
    """
    Apply styling to the total present table with improved UI
    """
    def highlight_top_performers(row):
        max_present = df['Total Present'].max()
        if row['Total Present'] == max_present:
            return ['background-color: #E8F5E8; color: #000000; font-weight: bold'] * len(row)
        elif row['Total Present'] >= max_present * 0.8:
            return ['background-color: #F0F8FF; color: #000000; font-weight: bold'] * len(row)
        return [''] * len(row)
   
    styled_df = df.style.apply(highlight_top_performers, axis=1)
    return styled_df

def create_excel_with_styling(df, calendar_df, employee_summary, total_present_summary, visualizations):
    """
    Create an Excel file with color coding and visualizations
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write dataframes to Excel
        df.to_excel(writer, sheet_name='Classified_Attendance', index=False)
        calendar_df.to_excel(writer, sheet_name='Monthly_Calendar', index=False)
        employee_summary.to_excel(writer, sheet_name='Employee_Summary', index=False)
        total_present_summary.to_excel(writer, sheet_name='Total_Present_Summary', index=False)
        
        # Get workbook and worksheets
        workbook = writer.book
        classified_sheet = writer.sheets['Classified_Attendance']
        calendar_sheet = writer.sheets['Monthly_Calendar']
        summary_sheet = writer.sheets['Employee_Summary']
        total_present_sheet = writer.sheets['Total_Present_Summary']
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply styling to Classified Attendance sheet
        for col in range(1, classified_sheet.max_column + 1):
            classified_sheet.cell(1, col).fill = header_fill
            classified_sheet.cell(1, col).font = header_font
            classified_sheet.cell(1, col).border = border
            
        # Color code attendance column in classified data
        if 'attendance' in df.columns:
            attendance_col_idx = df.columns.get_loc('attendance') + 1
            for row in range(2, classified_sheet.max_row + 1):
                cell_value = classified_sheet.cell(row, attendance_col_idx).value
                if cell_value in COLOR_MAP:
                    classified_sheet.cell(row, attendance_col_idx).fill = PatternFill(
                        start_color=COLOR_MAP[cell_value], end_color=COLOR_MAP[cell_value], fill_type='solid')
        
        # Apply styling to Calendar sheet
        for col in range(1, calendar_sheet.max_column + 1):
            calendar_sheet.cell(1, col).fill = header_fill
            calendar_sheet.cell(1, col).font = header_font
            calendar_sheet.cell(1, col).border = border
            
        # Color code calendar dates
        date_columns = [col for col in calendar_df.columns if any(day in str(col) for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'])]
        for col_name in date_columns:
            if col_name in calendar_df.columns:
                col_idx = calendar_df.columns.get_loc(col_name) + 1
                for row in range(2, calendar_sheet.max_row + 1):
                    cell_value = calendar_sheet.cell(row, col_idx).value
                    if cell_value in COLOR_MAP:
                        calendar_sheet.cell(row, col_idx).fill = PatternFill(
                            start_color=COLOR_MAP[cell_value], end_color=COLOR_MAP[cell_value], fill_type='solid')
        
        # Apply styling to Summary sheet
        for col in range(1, summary_sheet.max_column + 1):
            summary_sheet.cell(1, col).fill = header_fill
            summary_sheet.cell(1, col).font = header_font
            summary_sheet.cell(1, col).border = border
            
        # Highlight top performers in summary
        attendance_pct_col = employee_summary.columns.get_loc('Attendance %') + 1
        for row in range(2, summary_sheet.max_row + 1):
            attendance_pct = summary_sheet.cell(row, attendance_pct_col).value
            if isinstance(attendance_pct, (int, float)):
                if attendance_pct >= 90:
                    summary_sheet.cell(row, attendance_pct_col).fill = PatternFill(
                        start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif attendance_pct < 70:
                    summary_sheet.cell(row, attendance_pct_col).fill = PatternFill(
                        start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        # Apply styling to Total Present sheet
        for col in range(1, total_present_sheet.max_column + 1):
            total_present_sheet.cell(1, col).fill = header_fill
            total_present_sheet.cell(1, col).font = header_font
            total_present_sheet.cell(1, col).border = border
            
        # Highlight top performers in total present
        total_present_col = total_present_summary.columns.get_loc('Total Present') + 1
        max_present = total_present_summary['Total Present'].max()
        for row in range(2, total_present_sheet.max_row + 1):
            present_count = total_present_sheet.cell(row, total_present_col).value
            if present_count == max_present:
                for col in range(1, total_present_sheet.max_column + 1):
                    total_present_sheet.cell(row, col).fill = PatternFill(
                        start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        
    output.seek(0)
    return output

def load_data_from_file(file_path="./ATTENDANCE_DATA.xlsx"):
    """
    Load data from a local file if it exists
    """
    try:
        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            return df
        else:
            return None
    except Exception as e:
        st.error(f"Error loading local file: {str(e)}")
        return None

def main():
    st.set_page_config(page_title="Attendance Analytics", layout="wide")
   
    st.title("📊 Attendance Classification & Analytics Dashboard")
   
    # Create tabs
    tab1, tab2, tab3, tab4 = st.tabs(["📈 Dashboard", "📋 Data Analysis", "👤 Individual Analysis", "📅 Calendar View"])
   
    # File handling in sidebar
    with st.sidebar:
        st.header("Data Source")
        
        # Check if local file exists
        local_file_path = "./ATTENDANCE_DATA.xlsx"
        local_file_exists = os.path.exists(local_file_path)
        
        if local_file_exists:
            st.success(f"📁 Local file found: {local_file_path}")
            use_local_file = st.checkbox("Use local ATTENDANCE_DATA.xlsx", value=True)
        else:
            st.warning("📁 Local ATTENDANCE_DATA.xlsx not found")
            use_local_file = False
        
        # File upload option
        uploaded_file = None
        if not use_local_file:
            uploaded_file = st.file_uploader("Or upload your Excel file", type=['xlsx'])
            if uploaded_file is not None:
                st.success("File uploaded successfully!")
        
        st.header("Classification Rules")
        with st.expander("View Rules"):
            st.markdown("""
            - **Present**: Start time ≤ 9:15 AM AND Start meter < 200
            - **Outstation Present**: Present + outstation reason
            - **Late**: Start time > 9:15 AM OR Start meter ≥ 200
            - **Outstation Late**: Late + outstation reason
            """)

    # Load data based on selection
    df = None
    if use_local_file and local_file_exists:
        df = load_data_from_file(local_file_path)
        if df is not None:
            st.sidebar.success("✅ Using local ATTENDANCE_DATA.xlsx")
    elif uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            st.sidebar.success("✅ Using uploaded file")
        except Exception as e:
            st.error(f"Error reading uploaded file: {str(e)}")
    else:
        if not local_file_exists:
            st.info("👆 Please upload an Excel file to get started with attendance analysis.")
        else:
            st.info("👆 Select data source option to begin analysis.")

    if df is not None and not df.empty:
        try:
            # Process data
            df['attendance'] = df.apply(classify_attendance, axis=1)
            employee_summary = create_employee_summary(df)
            total_present_summary = create_total_present_summary(df)
            calendar_df = create_monthly_calendar_view(df)
            dashboard_metrics = create_dashboard_metrics(df, employee_summary)
            visualizations = create_visualizations(df, employee_summary, dashboard_metrics)
           
            # Dashboard Tab
            with tab1:
                st.header("🏢 Attendance Dashboard")
               
                # Key Metrics
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Total Employees", dashboard_metrics['total_employees'])
                with col2:
                    st.metric("Total Present Days", dashboard_metrics['total_present_days'])
                with col3:
                    st.metric("Total Late Days", dashboard_metrics['total_late_days'])
                with col4:
                    st.metric("Total Absent Days", dashboard_metrics['total_absent_days'])
                with col5:
                    st.metric("Overall Attendance Rate", f"{dashboard_metrics['overall_attendance_rate']}%")
               
                # First row of charts
                col1, col2 = st.columns(2)
                with col1:
                    st.plotly_chart(visualizations['attendance_pie'], use_container_width=True)
                with col2:
                    st.plotly_chart(visualizations['performance_bar'], use_container_width=True)
               
                # Second row of charts
                col1, col2 = st.columns(2)
                with col1:
                    st.plotly_chart(visualizations['top_employees'], use_container_width=True)
                with col2:
                    st.plotly_chart(visualizations['attendance_scatter'], use_container_width=True)
               
                # Third row - full width
                st.plotly_chart(visualizations['daily_trends'], use_container_width=True)
               
                # Quick Stats
                st.subheader("📊 Quick Statistics")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.info(f"**Best Attendance:** {employee_summary.loc[employee_summary['Attendance %'].idxmax(), 'User Name']} ({employee_summary['Attendance %'].max()}%)")
                with col2:
                    st.warning(f"**Most Late Arrivals:** {employee_summary.loc[employee_summary['Late'].idxmax(), 'User Name']} ({employee_summary['Late'].max()} days)")
                with col3:
                    st.error(f"**Most Absences:** {employee_summary.loc[employee_summary['Absent'].idxmax(), 'User Name']} ({employee_summary['Absent'].max()} days)")
           
            # Data Analysis Tab
            with tab2:
                st.header("📋 Detailed Data Analysis")
               
                st.subheader("Classified Data")
                st.dataframe(df[['User Name', 'User Code', 'Start Day Time',
                               'Start DiffIn Meters', 'Attendance Reason Start',
                               'Attendance Reason End', 'attendance']].head(), use_container_width=True)
               
                st.subheader("👥 Employee Attendance Summary")
                st.dataframe(style_summary(employee_summary), use_container_width=True)
               
                st.subheader("🏆 Total Present Days Summary")
                st.dataframe(style_total_present(total_present_summary), use_container_width=True)
               
                # Overall statistics
                st.subheader("Overall Attendance Summary")
                attendance_counts = df['attendance'].value_counts()
                st.write(attendance_counts)
               
                st.subheader("Attendance Distribution")
                st.bar_chart(attendance_counts)
           
            # Individual Analysis Tab
            with tab3:
                st.header("👤 Individual Employee Analysis")
               
                # Employee selection
                col1, col2 = st.columns([1, 3])
                with col1:
                    # Create combined list for selection
                    employee_options = ["Select Employee"] + \
                                     employee_summary['User Name'].tolist() + \
                                     employee_summary['User Code'].tolist()
                    selected_employee = st.selectbox(
                        "Select Employee:",
                        options=employee_options,
                        key="employee_selector"
                    )
               
                if selected_employee != "Select Employee":
                    # Get individual analysis
                    individual_analysis = create_individual_analysis(df, employee_summary, selected_employee)
                   
                    if individual_analysis:
                        user_name = individual_analysis['user_name']
                        user_code = individual_analysis['user_code']
                        summary = individual_analysis['summary']
                        detailed_records = individual_analysis['detailed_records']
                        indiv_visualizations = individual_analysis['visualizations']
                       
                        # Employee Header
                        st.subheader(f"Employee: {user_name} ({user_code})")
                       
                        # Key Metrics
                        col1, col2, col3, col4, col5, col6 = st.columns(6)
                        with col1:
                            st.metric("Present", summary['Present'])
                        with col2:
                            st.metric("Outstation Present", summary['Outstation Present'])
                        with col3:
                            st.metric("Total Present", summary['Total Present'])
                        with col4:
                            st.metric("Late", summary['Late'])
                        with col5:
                            st.metric("Outstation Late", summary['Outstation Late'])
                        with col6:
                            st.metric("Attendance %", f"{summary['Attendance %']}%")
                       
                        # Visualizations
                        st.subheader("📊 Individual Analytics")
                       
                        # First row of charts
                        col1, col2 = st.columns(2)
                        with col1:
                            st.plotly_chart(indiv_visualizations['individual_pie'], use_container_width=True)
                        with col2:
                            st.plotly_chart(indiv_visualizations['day_pattern'], use_container_width=True)
                       
                        # Second row of charts
                        col1, col2 = st.columns(2)
                        with col1:
                            st.plotly_chart(indiv_visualizations['weekly_trend'], use_container_width=True)
                        with col2:
                            if 'time_analysis' in indiv_visualizations:
                                st.plotly_chart(indiv_visualizations['time_analysis'], use_container_width=True)
                       
                        # Detailed Records Table
                        st.subheader("📋 Detailed Daily Records")
                        display_columns = ['date', 'day_name', 'Start Day Time', 'Start DiffIn Meters',
                                         'Attendance Reason Start', 'Attendance Reason End', 'attendance']
                        available_columns = [col for col in display_columns if col in detailed_records.columns]
                       
                        st.dataframe(
                            detailed_records[available_columns].rename(columns={
                                'date': 'Date',
                                'day_name': 'Day',
                                'Start Day Time': 'Start Time',
                                'Start DiffIn Meters': 'Start Diff (Meters)',
                                'Attendance Reason Start': 'Reason Start',
                                'Attendance Reason End': 'Reason End',
                                'attendance': 'Attendance'
                            }),
                            use_container_width=True
                        )
               
                else:
                    st.info("👆 Please select an employee from the dropdown to view individual analysis")
           
            # Calendar View Tab
            with tab4:
                st.header("📅 Monthly Calendar View")
               
                # Display calendar with styling
                st.write("**Color Legend:**")
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.markdown("🟩 **Present**")
                with col2:
                    st.markdown("🟦 **Outstation Present**")
                with col3:
                    st.markdown("🟥 **Late**")
                with col4:
                    st.markdown("🟨 **Outstation Late**")
                with col5:
                    st.markdown("⬜ **Absent**")
               
                # Display the calendar
                st.dataframe(style_calendar(calendar_df), use_container_width=True, height=600)
               
                # Calendar statistics
                if not df.empty:
                    sample_date = pd.to_datetime(df['Attendance Date'].iloc[0])
                    month_name = sample_date.strftime("%B %Y")
                    st.write(f"**Month:** {month_name}")
                   
                    total_days = len([col for col in calendar_df.columns if any(day in col for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'])])
                    total_users = len(calendar_df)
                    st.write(f"**Calendar Statistics:** {total_users} users × {total_days} days")
           
            # Download options in sidebar
            with st.sidebar:
                st.header("Download Reports")
               
                # Create enhanced Excel with styling and visualizations
                styled_excel = create_excel_with_styling(df, calendar_df, employee_summary, total_present_summary, visualizations)
               
                st.download_button(
                    label="📥 Download Full Report (Styled)",
                    data=styled_excel,
                    file_name="attendance_full_report_styled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
               
                # Also provide regular download option
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Classified_Attendance')
                    calendar_df.to_excel(writer, index=False, sheet_name='Monthly_Calendar')
                    employee_summary.to_excel(writer, index=False, sheet_name='Employee_Summary')
                    total_present_summary.to_excel(writer, index=False, sheet_name='Total_Present_Summary')
               
                output.seek(0)
               
                st.download_button(
                    label="📥 Download Basic Report",
                    data=output,
                    file_name="attendance_basic_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
               
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            st.error("Please make sure the file has the correct format with required columns.")
   
    else:
        # Welcome message when no file is available
        if not local_file_exists and uploaded_file is None:
            st.info("👆 Please upload an Excel file to get started with attendance analysis.")
       
        # Sample dashboard layout
        with tab1:
            st.header("Welcome to Attendance Analytics Dashboard")
            col1, col2 = st.columns(2)
            with col1:
                st.write("Once you upload your data, you'll see:")
                st.write("✅ Key metrics and KPIs")
                st.write("✅ Interactive charts and visualizations")
                st.write("✅ Employee performance analysis")
                st.write("✅ Daily attendance trends")
                st.write("✅ Individual employee analysis")
            with col2:
                st.write("Features include:")
                st.write("📊 Comprehensive dashboard")
                st.write("📋 Detailed data analysis")
                st.write("👤 Individual employee insights")
                st.write("📅 Monthly calendar view")
                st.write("📥 Exportable reports with color coding and charts")

if __name__ == "__main__":
    main()